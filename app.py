from flask import Flask, render_template, request, flash
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import os
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail

SENDGRID_API_KEY = "CEAGC9CENSYM41KL9BRJQ1YA"  # In production, set this as an environment variable

app = Flask(__name__)
app.secret_key = 'CEAGC9CENSYM41KL9BRJQ1YA'  # Change to something secret

@app.route('/')
def home():
    return render_template('home.html')

@app.route('/skills')
def skills():
    return render_template('skills.html')

@app.route('/education')
def education():
    return render_template('education.html')

@app.route('/resume')
def resume():
    return render_template('resume.html')

@app.route('/internships')
def internships():
    return render_template('internships.html')

@app.route('/projects')
def projects():
    return render_template('projects.html')

@app.route('/certifications')
def certifications():
    return render_template('certifications.html')

@app.route('/contact', methods=['GET', 'POST'])
def contact():
    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        feedback = request.form['feedback'].replace('\r\n', '\n').replace('\r', '\n')

        # Save feedback to Excel
        try:
            excel_path = "feedback.xlsx"
            if os.path.exists(excel_path):
                wb = load_workbook(excel_path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.append(['Name', 'Email', 'Feedback'])  # header row

            ws.append([name, email, feedback])
            for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True)
            wb.save(excel_path)
        except Exception as excel_e:
            flash(f"Failed to save feedback in Excel: {excel_e}", "danger")
            print(f"Excel error: {excel_e}")
            return render_template('contact.html')

        # Send feedback email using SendGrid
        try:
            message = Mail(
                from_email='iamviswanadhveera@gmail.com',  # Set an email verified on SendGrid
                to_emails='iamviswanadhveera@gmail.com,
                subject='Portfolio Feedback',
                html_content=f"<b>Name:</b> {name}<br><b>Email:</b> {email}<br><b>Feedback:</b><br>{feedback.replace(chr(10), '<br>')}"
            )
            sg = SendGridAPIClient(SENDGRID_API_KEY)
            sg.send(message)
            flash('Feedback sent and stored successfully!', 'success')
        except Exception as email_e:
            flash(f"Feedback saved, but email failed (SendGrid): {email_e}", "warning")
            print(f"SendGrid error: {email_e}")

    return render_template('contact.html')

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
